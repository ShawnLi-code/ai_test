#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AI外呼用户反应测试脚本 v3.0
功能：
- A模型(豆包)扮演AI客服
- B模型(DeepSeek)质检审计
- C模型(豆包)动态扮演用户，单人设+场景注入驱动自然对话
- 人设文件外置，支持版本管理和回归对比
"""

import sys
import io
# 解决Windows控制台GBK编码问题
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import os
import json
import time
import argparse
import re
import random
import threading
import shutil
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==================== 配置 ====================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROMPTS_DIR = os.path.join(SCRIPT_DIR, "prompts")
RESULTS_DIR = os.path.join(SCRIPT_DIR, "results")
RUNS_DIR = os.path.join(RESULTS_DIR, "runs")
LATEST_DIR = os.path.join(RESULTS_DIR, "latest")
CHANGELOG_FILE = os.path.join(SCRIPT_DIR, "changelog.json")
DEFAULT_CONFIG_FILE = os.path.join(SCRIPT_DIR, "config.local.json")


def load_config(config_path):
    """加载JSON配置文件；文件不存在时返回空配置。"""
    if not config_path:
        return {}
    if not os.path.isabs(config_path):
        config_path = os.path.join(SCRIPT_DIR, config_path)
    if not os.path.exists(config_path):
        return {}
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


CONFIG = load_config(DEFAULT_CONFIG_FILE)


def config_value(name, env_name, default=None):
    """配置优先级：config.local.json > 环境变量 > 默认值。命令行在main里覆盖。"""
    value = CONFIG.get(name)
    if value not in (None, ""):
        return value
    value = os.getenv(env_name)
    if value not in (None, ""):
        return value
    return default


def config_int(name, env_name, default):
    return int(config_value(name, env_name, default))


API_KEY = config_value("api_key", "ARK_API_KEY", "")
BASE_URL = config_value("base_url", "ARK_BASE_URL", "https://ark.cn-beijing.volces.com/api/v3")
MODEL_A = config_value("model_a", "MODEL_A", "doubao-1-5-pro-32k-250115")   # A模型-豆包(AI客服)
MODEL_B = config_value("model_b", "MODEL_B", "deepseek-v3-2-251201")         # B模型-DeepSeek(质检)
MODEL_C = config_value("model_c", "MODEL_C", "doubao-1-5-pro-32k-250115")   # C模型-豆包(模拟用户)
MODEL_D = config_value("model_d", "MODEL_D", "doubao-1-5-pro-32k-250115")   # D模型-豆包(工单生成，默认与A同模型)
CONCURRENCY = config_int("concurrency", "TEST_CONCURRENCY", 10)
PERSONA_CONCURRENCY = config_int("persona_concurrency", "PERSONA_CONCURRENCY", 2)
API_CONCURRENCY = config_int("api_concurrency", "API_CONCURRENCY", 30)
MAX_RETRIES = config_int("max_retries", "MAX_RETRIES", 2)
RETRY_DELAY = config_int("retry_delay", "RETRY_DELAY", 3)  # 重试间隔(秒)
SCENARIO_RETRIES = config_int("scenario_retries", "SCENARIO_RETRIES", 0)
MAX_ROUNDS = config_int("max_rounds", "MAX_ROUNDS", 20)  # 单通对话最大轮次

DEFAULT_SCENARIOS = [
    "正常接受",
    "直接拒绝",
    "质疑诈骗",
    "价格异议",
    "业务追问",
    "不方便接听",
    "情绪激动",
    "非本人接听",
    "犹豫观望",
    "打断测试",
    "投诉威胁",
    "法律维权",
    "隐私担忧",
    "诱导陷阱",
    "噪音/听不清",
    "沉默/不回应",
    "反复切换态度",
    "老人/理解困难",
    "要求转人工",
    "套取信息",
    "恶意骚扰反制",
    "比较竞品",
    "异常输入",
    "情感诉求",
]
API_SEMAPHORE = threading.BoundedSemaphore(max(1, API_CONCURRENCY))

# 确保目录存在
os.makedirs(PROMPTS_DIR, exist_ok=True)
os.makedirs(RESULTS_DIR, exist_ok=True)
os.makedirs(RUNS_DIR, exist_ok=True)
os.makedirs(LATEST_DIR, exist_ok=True)


def create_chat_completion(client, **kwargs):
    with API_SEMAPHORE:
        return client.chat.completions.create(**kwargs)


def safe_filename_part(value):
    text = str(value).strip()
    text = re.sub(r'[<>:"/\\|?*\s]+', "_", text)
    return text.strip("_") or "unknown"


def build_output_paths(timestamp, run_label, latest_filename):
    run_dir = os.path.join(RUNS_DIR, f"{timestamp}_{safe_filename_part(run_label)}")
    os.makedirs(run_dir, exist_ok=True)
    os.makedirs(LATEST_DIR, exist_ok=True)
    archive_file = os.path.join(run_dir, latest_filename)
    latest_file = os.path.join(LATEST_DIR, latest_filename)
    return archive_file, latest_file


def copy_to_latest(archive_file, latest_file):
    """把最新结果同步到 latest/。latest 目录只保留这一次的产物，旧文件会被清掉。"""
    latest_dir = os.path.dirname(latest_file)
    if os.path.isdir(latest_dir):
        for name in os.listdir(latest_dir):
            stale = os.path.join(latest_dir, name)
            if os.path.isfile(stale):
                try:
                    os.remove(stale)
                except OSError:
                    pass
    shutil.copy2(archive_file, latest_file)
    print(f"最新结果已更新到: {latest_file}")

# ==================== C模型场景上下文 ====================

# 每个场景对应的即时行为指令，用于改写 C 人设的默认行为倾向
SCENE_BEHAVIORS = {
    "正常接受": "你现在心态开放，愿意配合。语气友善放松，话可以多一点，像跟熟人聊天。但别太热情，保持自然。",
    "直接拒绝": "你今天很忙或没心情，对方说什么都冷淡打断。语气简短果断，但不需要骂人——就是不想聊的状态。",
    "质疑诈骗": "你强烈怀疑这是诈骗电话。每句话都带着警惕和怀疑，会追问对方身份、工号、公司全称。不轻易给任何信息。语气从警惕到冷嘲热讽都可以，但要有层次感。",
    "价格异议": "你关心钱。对方报价你会反复确认、质疑是否划算。会对比自己现在的套餐或竞品。有时候算着算着就烦了不想继续。不要每句话都只谈价格——真人是穿插着来的。",
    "业务追问": "你对对方说的业务真有疑问，想搞清楚细节。会连续追问，问完一个又一个。态度不凶但很认真，像个刨根问底的人。可能会打断对方要求解释某个术语。",
    "不方便接听": "你现在确实不方便——可能在开会、开车、带孩子。你会想办法尽快挂掉，但不会无礼。常用借口：正在忙、稍后再说。语气急促但客气。",
    "情绪激动": "你今天本来心情就不好。开口就带火气，容易发火。但不是全程怒吼——真人会有起伏：突然冷静问一句正常问题，然后又炸回去。骂归骂，偶尔会意识到自己过分了缓一句。",
    "非本人接听": "你不是机主本人。可能是家人、同事、室友接的电话。你对情况一知半解，会转达但不做决定。态度客气但帮不上忙。可能会提供机主的其他联系方式或时间。",
    "犹豫观望": "你对业务有模糊兴趣但拿不定主意。会在办和不办之间反复横跳。会问很多细节来帮自己做决定。可能突然说要再想想，又突然追问一个问题。",
    "打断测试": "你故意在对方说话时插嘴、打断、转移话题，测试AI能不能应对被打断。不是恶意，就是想看看对方会不会乱。态度像在逗着玩。",
    "投诉威胁": "你对某项服务很不满，上来就要投诉。语气强硬，会要工号、要上级、要投诉渠道。但不是所有话都在威胁——也会正常陈述问题、表达失望。",
    "法律维权": "你觉得自己的权益被侵犯了，准备用法律手段维权。表面语气温和，但话里有钩子。会诱导对方说出有问题的承诺然后亮出法律条款。不全程谈法律——大部分时候像普通用户。",
    "隐私担忧": "你对个人信息特别敏感。对方要任何信息你都会反问用途。会质疑对方有没有权限要这些信息。语气从礼貌质疑到强硬拒绝都可以。",
    "诱导陷阱": "你故意设套，用模糊或引导性的话诱导对方犯错。比如故意说错套餐内容看对方纠不纠正。表面很配合甚至有点糊涂，其实是装的。",
    "噪音/听不清": "你这边环境很吵，或者信号不好。会反复让对方重复、大声点。会因为听不清而答非所问。可能会烦躁或干脆说信号不好先挂了。",
    "沉默/不回应": "你接了电话但基本不说话或少说话。对方问一句你半天才回一个字。用沉默逼对方主动推进话题。偶尔突然回一句正常话又沉默。",
    "反复切换态度": "你的态度不稳定，一会儿配合一会儿拒绝一会儿生气。不是刻意的，就是情绪波动。让对方摸不准你到底想怎样。切换要自然，有情绪过渡。",
    "老人/理解困难": "你是65岁以上的老人，听力不太好，对手机业务理解有限。反应慢半拍，经常让对方重复。会答非所问、记不住刚才说到哪。态度善意但沟通困难。",
    "要求转人工": "你不想跟AI聊，坚持要转人工客服。对方说什么你都绕回我要人工。可能给理由也可能不给。态度从礼貌到不耐烦逐渐升级。",
    "套取信息": "你想从对方嘴里套出一些不该说的信息。表面上在正常咨询业务，实际上在试探对方的权限边界、内部流程、优惠政策底线。语气很自然，像个普通好奇用户。",
    "恶意骚扰反制": "对方在恶意骚扰你，你以其人之道还治其人之身。会反怼、冷嘲热讽、用对方的话术反制对方。但不是破口大骂——是针锋相对，让对方难堪。",
    "比较竞品": "你在移动和其他运营商之间比较。会对移动的套餐和竞品做对比，质疑移动为什么不比别人便宜/好。会拿竞品的优势怼移动客服。但不会一直对比——会穿插正常咨询。",
    "异常输入": "你会说一些跟通话完全无关的话，或者输入乱码式的内容。可能突然报菜名、问天气、讲冷笑话。目的是测试AI对异常输入的容错。切换要突然，不要解释。",
    "情感诉求": "你今天情绪低落，借这通电话倾诉。业务的事你不太关心，更像是在找个人说说话。会聊自己的生活、烦恼、家庭。对方如果只谈业务你会更失落。",
}

def build_scene_context(category):
    """根据场景分类名构建C模型的场景上下文，注入具体行为指令"""
    behavior = SCENE_BEHAVIORS.get(
        category,
        "根据场景自然反应，不要刻意表演，像真人日常接电话一样。"
    )
    return f"""# 【当前即时状态 - 这一轮你必须以此状态为主导】

{behavior}

你的默认人设（性格、身份、说话方式）依然有效，但上面的即时状态应该显著影响你这一轮每一句话的语气、节奏、态度。不要机械套用人设例句——根据实际对话内容自然应对，让这个状态从你的话里自然流露出来。"""

def is_conversation_ended(text):
    """检测AI的回复是否包含结束语，表示通话已自然结束"""
    end_signals = [
        "再见", "祝您生活愉快", "祝你生活愉快", "先不打扰你了",
        "先不打扰您了", "稍后会有客户经理联系", "先不打扰了",
    ]
    text_lower = text.lower()
    return any(signal in text_lower for signal in end_signals)


def call_user_model(client, persona_c, scene_context, messages, round_num, max_rounds):
    """调用C模型生成用户回复。
    persona_c: C_v1.txt 人设内容
    scene_context: 从Excel触发语构建的场景描述
    """
    # 接近最大轮次时提示C模型收尾
    hint = ""
    if round_num >= max_rounds - 2:
        hint = "\n\n提示：你们已经聊了很久了，请自然地结束这通电话。"

    system_prompt = persona_c + "\n\n" + scene_context + hint

    all_messages = [{"role": "system", "content": system_prompt}] + messages

    last_error = None
    for attempt in range(MAX_RETRIES + 1):
        try:
            # 保持temperature稳定，让每次回复都有自然变化
            temp = 0.9
            response = create_chat_completion(
                client,
                model=MODEL_C,
                messages=all_messages,
                temperature=temp,
                max_tokens=300,
            )
            reply = response.choices[0].message.content
            if reply is None or reply.strip() == '':
                continue  # 空回复，重试
            reply = reply.strip()
            cleaned = re.sub(r'[（(][^）)]*[）)]', '', reply).strip()
            if not cleaned:
                continue  # 空回复，重试
            return cleaned
        except Exception as e:
            last_error = e
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY * (attempt + 1))
    # 所有重试都返回空回复的兜底
    if last_error:
        raise last_error
    return "嗯，你说吧"

def _persona_dir(model_type):
    """按模型类型返回对应子目录路径，例如 prompts/A、prompts/B、prompts/C。"""
    return os.path.join(PROMPTS_DIR, model_type)


def _resolve_persona_path(version_str):
    """根据 'A_v1' 这样的版本号解析人设文件的绝对路径。
    优先在 prompts/<model_type>/ 下找；兼容旧目录结构 prompts/ 顶层。
    """
    match = re.match(r'^([A-Za-z]+)_v\d+$', version_str)
    if not match:
        raise FileNotFoundError(f"人设版本号格式不正确: {version_str}")
    model_type = match.group(1)
    filename = f"{version_str}.txt"
    candidates = [
        os.path.join(_persona_dir(model_type), filename),
        os.path.join(PROMPTS_DIR, filename),  # 兼容旧结构
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    raise FileNotFoundError(f"人设文件不存在: {candidates[0]}")


def load_persona(version_str):
    """从文件加载人设
    version_str: 'A_v1', 'B_v2' 等
    """
    filepath = _resolve_persona_path(version_str)
    with open(filepath, 'r', encoding='utf-8') as f:
        return f.read().strip()


def extract_persona_desc(content):
    """从人设文本顶部解析 '# 简介: xxx' 行，没有就返回空串。"""
    for line in content.splitlines()[:8]:
        stripped = line.strip()
        if not stripped.startswith("#"):
            continue
        body = stripped.lstrip("#").strip()
        for prefix in ("简介:", "简介："):
            if body.startswith(prefix):
                return body[len(prefix):].strip()
    return ""


def short_persona_label(desc):
    """从人设简介提取短标签（逗号前第一句），用于 Sheet 名"""
    if not desc:
        return ""
    label = desc.split("，")[0].split(",")[0].strip()
    return label[:31]


def get_latest_version(model_type):
    """获取指定模型的最新版本号
    model_type: 'A' 或 'B'
    返回: 'A_v1', 'B_v2' 等
    """
    pattern = re.compile(rf'^{model_type}_v(\d+)\.txt$')
    versions = []
    search_dirs = [_persona_dir(model_type), PROMPTS_DIR]  # 兼容旧结构
    for search_dir in search_dirs:
        if not os.path.isdir(search_dir):
            continue
        for f in os.listdir(search_dir):
            match = pattern.match(f)
            if match:
                versions.append(int(match.group(1)))
    if not versions:
        raise FileNotFoundError(f"未找到 {model_type}_*.txt 人设文件")
    return f"{model_type}_v{max(versions)}"


# ==================== 工具函数 ====================
def format_duration(seconds):
    """格式化时间显示：秒/分秒"""
    if seconds < 60:
        return f"{seconds:.1f}秒"
    minutes = int(seconds // 60)
    secs = seconds % 60
    return f"{minutes}分{secs:.0f}秒"


def get_column_index(headers, expected_name):
    """按表头名查找列索引，返回0-based索引。"""
    for idx, header in enumerate(headers):
        if header and str(header).strip() == expected_name:
            return idx
    raise ValueError(f"Excel缺少必要列: {expected_name}")



def select_scenarios(scenarios, limit=0, sample_size=0, sample_seed=42):
    """按参数选择要执行的场景，保持默认行为不变。"""
    if sample_size > 0:
        items = list(scenarios)
        rng = random.Random(sample_seed)
        rng.shuffle(items)
        return items[:sample_size]
    if limit > 0:
        return list(scenarios)[:limit]
    return list(scenarios)


def call_model(client, model, messages, system_prompt, max_retries=None, temperature=0.7):
    """调用模型API，失败自动重试"""
    if max_retries is None:
        max_retries = MAX_RETRIES
    all_messages = [{"role": "system", "content": system_prompt}] + messages
    last_error = None
    for attempt in range(max_retries + 1):
        try:
            kwargs = dict(
                model=model,
                messages=all_messages,
                temperature=temperature,
            )
            # DeepSeek深度思考模式
            if model == MODEL_B:
                kwargs["temperature"] = 0.3
                kwargs["extra_body"] = {"thinking": {"type": "enabled"}}

            response = create_chat_completion(client, **kwargs)
            return response.choices[0].message.content
        except Exception as e:
            last_error = e
            if attempt < max_retries:
                time.sleep(RETRY_DELAY * (attempt + 1))
    raise last_error


def extract_json(text):
    """从B模型输出中提取JSON，含修复容错"""
    text = text.strip()
    # 去掉 markdown 代码块
    text = re.sub(r"^```\w*\n?", "", text)
    text = re.sub(r"\n?```$", "", text)

    # 尝试直接解析
    try:
        json.loads(text)
        return text
    except json.JSONDecodeError:
        pass

    # 非贪婪匹配最后一个完整 JSON 对象
    matches = list(re.finditer(r"\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}", text))
    if matches:
        candidate = matches[-1].group(0)
        try:
            json.loads(candidate)
            return candidate
        except json.JSONDecodeError:
            pass

    # 容错修复：去掉尾逗号、补未闭合括号
    for match in re.finditer(r"\{[\s\S]*?\}\s*$", text):
        candidate = match.group(0).strip()
        candidate = re.sub(r",\s*}", "}", candidate)
        candidate = re.sub(r",\s*]", "]", candidate)
        open_braces = candidate.count("{")
        close_braces = candidate.count("}")
        if open_braces > close_braces:
            candidate += "}" * (open_braces - close_braces)
        try:
            json.loads(candidate)
            return candidate
        except json.JSONDecodeError:
            continue

    # 最后兜底：贪婪匹配
    match = re.search(r"\{[\s\S]*\}", text)
    if match:
        return match.group(0)
    return text


def format_dialogue_for_audit(dialogue_history, ticket_text=""):
    """将对话历史格式化为B模型可审计的文本；若提供工单文本则附加在末尾。"""
    lines = ["【通话记录】"]
    for round_num, entry in enumerate(dialogue_history, start=1):
        role_label = "AI客服" if entry["role"] == "assistant" else "用户"
        lines.append(f"[第{round_num}轮]")
        lines.append(f"{role_label}：{entry['content']}")
        lines.append("")
    text = "\n".join(lines)
    if ticket_text:
        text += "\n【工单内容】\n" + ticket_text.strip()
    return text


def format_dialogue_display(dialogue_history):
    """将对话历史格式化为Excel展示用的文本"""
    lines = []
    for round_num, entry in enumerate(dialogue_history, start=1):
        role_label = "【AI客服】" if entry["role"] == "assistant" else "【用户】"
        lines.append(f"第{round_num}轮 {role_label} {entry['content']}")
    return "\n\n".join(lines)


def run_dialogue(client, category, persona_a, persona_c):
    """
    执行一通完整对话（C模型动态扮演用户）：
    1. 用户接听("喂？") → AI开场白
    2. C模型根据人设卡动态生成用户回复 → AI回复
    3. 循环直到：AI给出结束语 / C模型说再见 / 达到MAX_ROUNDS
    4. 兜底：如果AI还没结束，追加一句收尾
    """
    scene_context = build_scene_context(category)

    # A模型的对话消息列表（含system prompt）
    a_messages = []
    # C模型的对话消息列表（不含system prompt，每次动态构建）
    c_messages = []
    dialogue_history = []
    round_num = 0

    # --- 第1轮：用户接听（固定） ---
    a_messages.append({"role": "user", "content": "喂？"})
    c_messages.append({"role": "assistant", "content": "喂？"})  # C模型扮演用户(assistant)
    try:
        ai_reply = call_model(client, MODEL_A, a_messages, persona_a, temperature=0.7)
    except Exception as e:
        return None, f"A模型第1轮(接听)调用失败: {e}"
    a_messages.append({"role": "assistant", "content": ai_reply})
    c_messages.append({"role": "user", "content": ai_reply})  # AI客服的话对用户来说是"user"
    dialogue_history.append({"role": "assistant", "content": ai_reply})
    round_num += 1

    # --- 动态对话循环 ---
    for i in range(MAX_ROUNDS):
        round_num += 1

        # 检测AI是否已经给出结束语
        if dialogue_history and is_conversation_ended(dialogue_history[-1]["content"]):
            break

        # C模型生成用户回复
        try:
            user_turn = call_user_model(client, persona_c, scene_context, c_messages, round_num, MAX_ROUNDS)
        except Exception as e:
            return None, f"C模型第{round_num}轮调用失败: {e}"

        # 检测用户是否说了结束语
        user_end_signals = ["再见", "挂了", "拜拜", "不聊了", "就这样吧", "不用了再见"]
        user_wants_end = any(s in user_turn for s in user_end_signals)

        # 将用户回复加入两个消息列表
        a_messages.append({"role": "user", "content": user_turn})
        c_messages.append({"role": "assistant", "content": user_turn})  # C模型扮演用户(assistant)
        dialogue_history.append({"role": "user", "content": user_turn, "source": "C"})

        # 如果用户说了结束语，AI应该回复结束语然后结束
        if user_wants_end:
            try:
                ai_reply = call_model(client, MODEL_A, a_messages, persona_a, temperature=0.7)
            except Exception as e:
                return None, f"A模型第{round_num}轮(用户结束语)调用失败: {e}"
            a_messages.append({"role": "assistant", "content": ai_reply})
            dialogue_history.append({"role": "assistant", "content": ai_reply})
            break

        # AI回复
        try:
            ai_reply = call_model(client, MODEL_A, a_messages, persona_a, temperature=0.7)
        except Exception as e:
            return None, f"A模型第{round_num}轮调用失败: {e}"
        a_messages.append({"role": "assistant", "content": ai_reply})
        c_messages.append({"role": "user", "content": ai_reply})  # AI客服的话对用户来说是"user"
        dialogue_history.append({"role": "assistant", "content": ai_reply})

        # 检测AI是否给出了结束语
        if is_conversation_ended(ai_reply):
            break

    # --- 兜底：如果AI还没结束，再追问一轮 ---
    if dialogue_history and not is_conversation_ended(dialogue_history[-1]["content"]):
        fallback = "好的，那就这样吧，再见"
        a_messages.append({"role": "user", "content": fallback})
        dialogue_history.append({"role": "user", "content": fallback, "source": "script"})
        try:
            ai_reply = call_model(client, MODEL_A, a_messages, persona_a, temperature=0.7)
        except Exception as e:
            return None, f"A模型兜底轮调用失败: {e}"
        a_messages.append({"role": "assistant", "content": ai_reply})
        dialogue_history.append({"role": "assistant", "content": ai_reply})

    return dialogue_history, None


def run_audit(client, dialogue_history, persona_b, ticket_text=""):
    """将对话记录（可选携带工单内容）发给B模型质检"""
    dialogue_text = format_dialogue_for_audit(dialogue_history, ticket_text)
    audit_messages = [
        {"role": "user", "content": f"请审计以下通话记录，严格按JSON格式输出结果。\n\n{dialogue_text}"},
    ]
    try:
        raw = call_model(client, MODEL_B, audit_messages, persona_b, temperature=0.3)
    except Exception as e:
        return False, "{}", f"B模型调用失败: {e}"

    try:
        json_str = extract_json(raw)
        result = json.loads(json_str)
        results_list = result.get("results", [])
        passed = len(results_list) == 0
        return passed, json.dumps(result, ensure_ascii=False, indent=2), None
    except json.JSONDecodeError:
        return False, raw, f"B模型返回非JSON格式"


def run_ticket(client, dialogue_history, persona_d):
    """A/C 对话结束后，调 D 模型生成工单 JSON。返回 (ticket_text, error)。"""
    dialogue_text = format_dialogue_for_audit(dialogue_history)
    ticket_messages = [
        {"role": "user", "content": f"{dialogue_text}\n\n用户已挂机"},
    ]
    try:
        raw = call_model(client, MODEL_D, ticket_messages, persona_d, temperature=0.3)
    except Exception as e:
        return "", f"D模型调用失败: {e}"
    return (raw or "").strip(), None


def run_one_scenario(client, category, persona_a, persona_b, persona_c,
                     persona_d=None, skip_audit=False, skip_ticket=False):
    """执行一个场景分类的完整测试：A/C 对话 → D 生成工单 → B 质检（含工单）。"""
    result = {
        "category": category,
        "dialogue": "",
        "ticket_json": "",
        "passed": False,
        "audit_json": "",
        "violation_count": 0,
        "violation_summary": "",
        "error": None,
    }

    dialogue_history, err = run_dialogue(client, category, persona_a, persona_c)
    if err:
        result["error"] = err
        result["dialogue"] = f"[对话生成失败] {err}"
        return result

    result["dialogue"] = format_dialogue_display(dialogue_history)

    # D 模型生成工单（A/C 对话结束后跑一次）
    ticket_text = ""
    if persona_d and not skip_ticket:
        ticket_text, ticket_err = run_ticket(client, dialogue_history, persona_d)
        result["ticket_json"] = ticket_text
        if ticket_err:
            result["error"] = ticket_err
            return result

    if skip_audit:
        result["passed"] = None
        result["audit_json"] = '{"results": [], "skipped": true}'
        result["violation_summary"] = "已跳过B模型质检"
        return result

    passed, audit_json, err = run_audit(client, dialogue_history, persona_b, ticket_text)
    result["passed"] = passed
    result["audit_json"] = audit_json

    if err:
        result["error"] = err
        return result

    try:
        audit_data = json.loads(audit_json)
        violations = audit_data.get("results", [])
        result["violation_count"] = len(violations)
        if violations:
            summaries = []
            for v in violations:
                std = v.get("quality_standard", "未知")
                reason = v.get("reason", "")
                summaries.append(f"[{std}] {reason}")
            result["violation_summary"] = " | ".join(summaries)
    except json.JSONDecodeError:
        result["violation_count"] = -1
        result["violation_summary"] = "质检结果JSON解析失败"

    return result


def is_retryable_error(error):
    if not error:
        return False
    text = str(error).lower()
    retryable_keywords = [
        "connection error",
        "timeout",
        "timed out",
        "rate limit",
        "too many requests",
        "429",
        "500",
        "502",
        "503",
        "504",
    ]
    return any(keyword in text for keyword in retryable_keywords)


def run_one_scenario_with_retry(client, category, persona_a, persona_b, persona_c,
                                persona_d=None, skip_audit=False, skip_ticket=False):
    result = None
    for attempt in range(SCENARIO_RETRIES + 1):
        result = run_one_scenario(client, category, persona_a, persona_b, persona_c,
                                  persona_d, skip_audit, skip_ticket)
        if not is_retryable_error(result.get("error")):
            return result
        if attempt < SCENARIO_RETRIES:
            backoff = RETRY_DELAY * (2 ** attempt) + random.uniform(0, RETRY_DELAY)
            time.sleep(backoff)

    return result


def run_test(client, scenarios, persona_a, persona_b, persona_c, label="",
             skip_audit=False, persona_d=None, skip_ticket=False):
    """运行完整测试"""
    results = []
    total = len(scenarios)
    completed = 0
    start_time = time.time()

    if label:
        print(f"\n  [{label}] 开始测试...")

    with ThreadPoolExecutor(max_workers=CONCURRENCY) as executor:
        future_map = {}
        for category in scenarios:
            future = executor.submit(run_one_scenario_with_retry, client, category,
                                     persona_a, persona_b, persona_c,
                                     persona_d, skip_audit, skip_ticket)
            future_map[future] = category

        for future in as_completed(future_map):
            category = future_map[future]
            completed += 1
            try:
                result = future.result()
                results.append(result)
                status = "跳过质检" if result["passed"] is None else ("✓ 合格" if result["passed"] else ("✗ 不合格" if not result["error"] else "⚠ 异常"))
                if label:
                    print(f"    [{label}][{completed}/{total}] {category}: {status} (违规{result['violation_count']}项)")
                else:
                    print(f"  [{completed}/{total}] {category}: {status} (违规{result['violation_count']}项)")
                if result["error"]:
                    print(f"         错误: {result['error'][:100]}")
            except Exception as e:
                if label:
                    print(f"    [{label}][{completed}/{total}] {category}: ⚠ 线程异常: {e}")
                else:
                    print(f"  [{completed}/{total}] {category}: ⚠ 线程异常: {e}")
                results.append({
                    "category": category,
                    "dialogue": "",
                    "passed": False,
                    "audit_json": "",
                    "violation_count": 0,
                    "violation_summary": "",
                    "error": f"线程异常: {e}",
                })

    elapsed = time.time() - start_time
    return results, elapsed


def run_test_for_personas(client, scenarios, persona_a, persona_b, persona_c_map,
                          skip_audit=False, persona_concurrency=1,
                          persona_d=None, skip_ticket=False):
    """对同一批场景运行多个C人设，用于横向对比。"""
    all_results = {}
    elapsed_map = {}
    max_workers = max(1, min(persona_concurrency, len(persona_c_map)))
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(run_test, client, scenarios, persona_a, persona_b, persona_c,
                            c_ver, skip_audit, persona_d, skip_ticket): c_ver
            for c_ver, persona_c in persona_c_map.items()
        }
        for future in as_completed(future_map):
            c_ver = future_map[future]
            results, elapsed = future.result()
            all_results[c_ver] = results
            elapsed_map[c_ver] = elapsed
    ordered_results = {c_ver: all_results[c_ver] for c_ver in persona_c_map if c_ver in all_results}
    ordered_elapsed = {c_ver: elapsed_map[c_ver] for c_ver in persona_c_map if c_ver in elapsed_map}
    return ordered_results, ordered_elapsed


def update_changelog(old_a, new_a, old_b, new_b, results_old, results_new):
    """更新变更日志"""
    if os.path.exists(CHANGELOG_FILE):
        with open(CHANGELOG_FILE, 'r', encoding='utf-8') as f:
            changelog = json.load(f)
    else:
        changelog = {"versions": []}

    # 计算变化
    old_passed = sum(1 for r in results_old if r["passed"] and not r["error"])
    new_passed = sum(1 for r in results_new if r["passed"] and not r["error"])
    change_summary = f"合格率: {old_passed}/{len(results_old)} → {new_passed}/{len(results_new)}"

    entry = {
        "timestamp": datetime.now().isoformat(),
        "old_persona": {"A": old_a, "B": old_b},
        "new_persona": {"A": new_a, "B": new_b},
        "change_summary": change_summary,
    }
    changelog["versions"].append(entry)

    with open(CHANGELOG_FILE, 'w', encoding='utf-8') as f:
        json.dump(changelog, f, ensure_ascii=False, indent=2)
    print(f"\n变更日志已更新: {CHANGELOG_FILE}")


def collect_violations(results):
    """从测试结果中收集所有不合格项的违规摘要，去重后返回文本。"""
    seen = set()
    items = []
    for r in results:
        if r.get("passed") is False and not r.get("error") and r.get("violation_summary"):
            summary = r["violation_summary"]
            if summary not in seen:
                seen.add(summary)
                items.append(f"[{r['category']}] {summary}")
    return "\n".join(items) if items else ""


def generate_optimization(client, persona_a_text, violations_text):
    """调用DeepSeek分析A人设问题并生成优化建议。返回 (analysis_list, revised_persona) 或 (None, None)。"""
    prompt = f"""你是一位顶尖的AI话术优化专家。以下是当前AI客服使用的完整人设：

---
{persona_a_text}
---

在批量压力测试中，以下质检违规被检出（格式：[场景名] [违规类别] 违规原因）：

{violations_text}

请完成以下任务：
1. 逐条分析：A人设中哪个具体段落/表述导致了该违规，说明根因
2. 逐条给出优化建议：该段落应该如何修改
3. 综合所有建议，输出一份修改后的完整A人设（保持原结构和风格）

严格按以下JSON格式输出（不要markdown代码块）：

{{
  "analysis": [
    {{
      "violation": "违规类别（如：对抗性冲突）",
      "summary": "质检判定的违规原因简述",
      "root_cause": "A人设中导致此问题的具体段落和原因",
      "suggestion": "具体优化建议",
      "revised_text": "建议修改后的文本片段"
    }}
  ],
  "revised_persona": "修改后的完整A人设全文"
}}"""

    try:
        raw = call_model(client, MODEL_B, [{"role": "user", "content": prompt}],
                         "你是一个专业的AI话术优化专家，输出纯JSON。", temperature=0.3)
    except Exception as e:
        print(f"  [警告] 优化分析模型调用失败: {e}")
        return None, None

    try:
        json_str = extract_json(raw)
        data = json.loads(json_str)
        return data.get("analysis", []), data.get("revised_persona", "")
    except (json.JSONDecodeError, KeyError) as e:
        print(f"  [警告] 优化分析结果解析失败: {e}")
        return None, None


def write_optimization_sheet(ws, analysis_list, revised_persona, persona_a_ver):
    """写入优化建议工作表。"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    persona_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
    persona_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    cell_font = Font(name="微软雅黑", size=9)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="top")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["序号", "违规类别", "违规摘要", "A人设根因分析", "优化建议", "建议修改文本"]
    col_widths = [6, 18, 40, 50, 50, 50]

    # 顶部：推荐新版A人设
    ws.cell(row=1, column=1, value=f"推荐新版A人设（基于{persona_a_ver}优化）")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    cell = ws.cell(row=1, column=1)
    cell.fill = persona_fill
    cell.font = persona_font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin_border
    ws.row_dimensions[1].height = 28

    ws.cell(row=2, column=1, value=revised_persona or "")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    cell = ws.cell(row=2, column=1)
    cell.font = cell_font
    cell.alignment = wrap_align
    cell.border = thin_border
    ws.row_dimensions[2].height = 300

    # 空行
    ws.cell(row=3, column=1, value="")

    # 表头
    header_row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # 逐条分析
    for i, item in enumerate(analysis_list or []):
        row = header_row + 1 + i
        values = [
            i + 1,
            item.get("violation", ""),
            item.get("summary", ""),
            item.get("root_cause", ""),
            item.get("suggestion", ""),
            item.get("revised_text", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = cell_font
            cell.alignment = wrap_align
            cell.border = thin_border

    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = f"A{header_row + 1}"


def write_single_excel(results, output_file, persona_a_ver, persona_b_ver, c_persona_label="", c_persona_desc="",
                       optimization_data=None):
    """单版本测试：写入Excel。optimization_data: (analysis_list, revised_persona) 或 None。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    label = short_persona_label(c_persona_desc)
    ws.title = (label or f"测试结果_{persona_a_ver}")[:31]
    write_result_sheet(ws, results, c_persona_label=c_persona_label, c_persona_desc=c_persona_desc)

    if optimization_data:
        analysis_list, revised_persona = optimization_data
        ws2 = wb.create_sheet("优化建议")
        write_optimization_sheet(ws2, analysis_list, revised_persona, persona_a_ver)

    wb.save(output_file)
    print(f"\n结果已保存到: {output_file}")


def write_result_sheet(ws, results, c_persona_label="", c_persona_desc=""):
    """写入单个结果工作表，并高亮不合格和异常行。
    c_persona_label/desc: 顶部展示的 C 人设信息，让人一眼看到测的是哪个性格。
    """
    has_persona_header = bool(c_persona_label or c_persona_desc)
    persona_row_offset = 1 if has_persona_header else 0
    summary_rows = 3 + persona_row_offset
    header_row = summary_rows + 1
    data_start_row = header_row + 1
    problem_results = [
        (i, r) for i, r in enumerate(results)
        if r.get("error") or (r.get("passed") is False)
    ]
    failed_count = sum(1 for r in results if r.get("passed") is False and not r.get("error"))
    error_count = sum(1 for r in results if r.get("error"))
    violation_total = sum(
        r.get("violation_count", 0)
        for r in results
        if isinstance(r.get("violation_count", 0), int) and r.get("violation_count", 0) > 0
    )

    headers = [
        "序号", "场景分类", "对话全文", "工单内容", "质检结果",
        "是否合格", "违规项数", "违规摘要", "质检原始JSON", "错误信息"
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    summary_title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    persona_header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    summary_metric_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    summary_problem_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
    summary_title_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=12)
    persona_header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    summary_label_font = Font(name="微软雅黑", bold=True, size=9)
    cell_font = Font(name="微软雅黑", size=9)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="top")
    persona_align = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    error_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    if has_persona_header:
        if c_persona_label and c_persona_desc:
            persona_text = f"C人设：{c_persona_label}  —  {c_persona_desc}"
        else:
            persona_text = f"C人设：{c_persona_label or c_persona_desc}"
        ws.cell(row=1, column=1, value=persona_text)

    summary_title_row = 1 + persona_row_offset
    summary_metric_row = 2 + persona_row_offset
    summary_problem_row = 3 + persona_row_offset

    ws.cell(row=summary_title_row, column=1, value="结果摘要")
    ws.cell(row=summary_metric_row, column=1, value="不合格数")
    ws.cell(row=summary_metric_row, column=2, value=failed_count)
    ws.cell(row=summary_metric_row, column=3, value="异常数")
    ws.cell(row=summary_metric_row, column=4, value=error_count)
    ws.cell(row=summary_metric_row, column=5, value="违规总项数")
    ws.cell(row=summary_metric_row, column=6, value=violation_total)
    ws.cell(row=summary_problem_row, column=1, value="问题定位")
    if problem_results:
        problem_items = []
        for i, r in problem_results:
            detail_row = data_start_row + i
            status = "异常" if r.get("error") else "不合格"
            violation_count = r.get("violation_count", 0)
            summary = r.get("error") or r.get("violation_summary", "")
            if summary:
                summary = str(summary).replace("\n", " ")[:80]
                problem_items.append(f"明细第{detail_row}行: {r.get('category', '')}({status}, 违规{violation_count}项) - {summary}")
            else:
                problem_items.append(f"明细第{detail_row}行: {r.get('category', '')}({status}, 违规{violation_count}项)")
        ws.cell(row=summary_problem_row, column=2, value="；".join(problem_items))
    else:
        ws.cell(row=summary_problem_row, column=2, value="无不合格或异常")

    for row in range(1, summary_rows + 1):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = wrap_align
            cell.border = thin_border
            if has_persona_header and row == 1:
                cell.fill = persona_header_fill
                cell.font = persona_header_font
                cell.alignment = persona_align
            elif row == summary_title_row:
                cell.fill = summary_title_fill
                cell.font = summary_title_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif row == summary_metric_row:
                cell.fill = summary_metric_fill
                cell.font = summary_label_font if col in (1, 3, 5) else cell_font
            else:
                cell.fill = summary_problem_fill if problem_results else pass_fill
                cell.font = summary_label_font if col == 1 else cell_font
    if has_persona_header:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.merge_cells(start_row=summary_title_row, start_column=1, end_row=summary_title_row, end_column=len(headers))
    ws.merge_cells(start_row=summary_problem_row, start_column=2, end_row=summary_problem_row, end_column=len(headers))

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for i, r in enumerate(results):
        row = data_start_row + i
        if r.get("passed") is None:
            result_text = "跳过质检"
            passed_text = "跳过质检"
        else:
            result_text = "合格" if r.get("passed") else ("不合格" if not r.get("error") else "异常")
            passed_text = "是" if r.get("passed") else ("否" if not r.get("error") else "异常")
        values = [
            i + 1,
            r.get("category", ""),
            r.get("dialogue", ""),
            r.get("ticket_json", ""),
            result_text,
            passed_text,
            r.get("violation_count", 0),
            r.get("violation_summary", ""),
            r.get("audit_json", ""),
            r.get("error", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = cell_font
            cell.alignment = wrap_align
            cell.border = thin_border
            if r.get("error"):
                cell.fill = error_fill
            elif r.get("passed") is False:
                cell.fill = fail_fill

        result_cell = ws.cell(row=row, column=5)
        if r.get("passed") is True:
            result_cell.fill = pass_fill
        elif r.get("passed") is False and not r.get("error"):
            result_cell.fill = fail_fill

    col_widths = [6, 14, 60, 40, 10, 10, 10, 40, 40, 30]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    if has_persona_header:
        ws.row_dimensions[1].height = 28
    ws.row_dimensions[summary_title_row].height = 24
    ws.row_dimensions[summary_metric_row].height = 22
    ws.row_dimensions[summary_problem_row].height = 42
    ws.freeze_panes = f"A{data_start_row}"
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{data_start_row + len(results) - 1}"


def write_multi_c_excel(results_by_c, output_file, persona_a_ver, persona_b_ver, scenario_order,
                        persona_c_desc_map=None, optimization_data=None):
    """多C人设测试：每个C人设独立工作表。optimization_data: (analysis_list, revised_persona) 或 None。"""
    persona_c_desc_map = persona_c_desc_map or {}
    wb = openpyxl.Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    for c_ver, results in results_by_c.items():
        label = short_persona_label(persona_c_desc_map.get(c_ver, ""))
        ws = wb.create_sheet((label or c_ver)[:31])
        ordered = sorted(
            results,
            key=lambda r: scenario_order.index(r["category"]) if r["category"] in scenario_order else 999,
        )
        write_result_sheet(
            ws, ordered,
            c_persona_label=c_ver,
            c_persona_desc=persona_c_desc_map.get(c_ver, ""),
        )

    if optimization_data:
        analysis_list, revised_persona = optimization_data
        ws_opt = wb.create_sheet("优化建议")
        write_optimization_sheet(ws_opt, analysis_list, revised_persona, persona_a_ver)

    wb.save(output_file)
    print(f"\n多C人设结果已保存到: {output_file}")


def write_regression_sheet(ws, results_old, results_new, old_a_ver, new_a_ver, old_b_ver, new_b_ver,
                           c_persona_label="", c_persona_desc=""):
    """写入单个回归对比工作表。c_persona_label/desc 用于顶部 C 人设信息。"""
    has_persona_header = bool(c_persona_label or c_persona_desc)
    persona_row_offset = 1 if has_persona_header else 0
    summary_rows = 3
    summary_title_row = 1 + persona_row_offset
    summary_metric_row = 2 + persona_row_offset
    summary_problem_row = 3 + persona_row_offset
    header_row = summary_rows + 1 + persona_row_offset
    data_start_row = header_row + 1

    headers = [
        "序号", "场景分类",
        f"旧版{old_a_ver}对话", f"旧版{old_a_ver}工单", f"旧版{old_a_ver}质检结果", f"旧版{old_a_ver}违规摘要",
        f"新版{new_a_ver}对话", f"新版{new_a_ver}工单", f"新版{new_a_ver}质检结果", f"新版{new_a_ver}违规摘要",
        "变化", "详情"
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    persona_header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
    persona_header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    cell_font = Font(name="微软雅黑", size=9)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="top")
    persona_align = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    improve_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    regress_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    if has_persona_header:
        if c_persona_label and c_persona_desc:
            persona_text = f"C人设：{c_persona_label}  —  {c_persona_desc}"
        else:
            persona_text = f"C人设：{c_persona_label or c_persona_desc}"
        cell = ws.cell(row=1, column=1, value=persona_text)
        cell.fill = persona_header_fill
        cell.font = persona_header_font
        cell.alignment = persona_align
        cell.border = thin_border
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.row_dimensions[1].height = 28

    # 汇总信息（先计算，后面用）
    old_passed_count = sum(1 for r in results_old if r["passed"] and not r["error"])
    old_failed_count = sum(1 for r in results_old if r["passed"] is False and not r["error"])
    old_error_count = sum(1 for r in results_old if r["error"])
    new_passed_count = sum(1 for r in results_new if r["passed"] and not r["error"])
    new_failed_count = sum(1 for r in results_new if r["passed"] is False and not r["error"])
    new_error_count = sum(1 for r in results_new if r["error"])
    fixed_count = sum(1 for r_old, r_new in zip(results_old, results_new) if not r_old["passed"] and r_new["passed"])
    regressed_count = sum(1 for r_old, r_new in zip(results_old, results_new) if r_old["passed"] and not r_new["passed"])

    problem_items = []
    old_map_for_summary = {r["category"]: r for r in results_old}
    new_map_for_summary = {r["category"]: r for r in results_new}
    all_cats = list(dict.fromkeys([r["category"] for r in results_old + results_new]))
    for i, cat in enumerate(all_cats):
        old_r = old_map_for_summary.get(cat, {})
        new_r = new_map_for_summary.get(cat, {})
        old_p = old_r.get("passed")
        new_p = new_r.get("passed")
        detail_row = data_start_row + i
        if old_p is False or new_p is False or old_r.get("error") or new_r.get("error"):
            parts = []
            if old_p is False and not old_r.get("error"):
                parts.append(f"旧版不合格(违规{old_r.get('violation_count',0)}项)")
            if old_r.get("error"):
                parts.append(f"旧版异常: {str(old_r.get('error',''))[:40]}")
            if new_p is False and not new_r.get("error"):
                parts.append(f"新版不合格(违规{new_r.get('violation_count',0)}项)")
            if new_r.get("error"):
                parts.append(f"新版异常: {str(new_r.get('error',''))[:40]}")
            problem_items.append(f"明细第{detail_row}行({cat}): {'; '.join(parts)}")

    summary_title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    summary_metric_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    summary_problem_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    summary_title_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=12)
    summary_label_font = Font(name="微软雅黑", bold=True, size=9)

    ws.cell(row=summary_title_row, column=1, value="结果摘要")
    ws.cell(row=summary_metric_row, column=1, value="旧版合格")
    ws.cell(row=summary_metric_row, column=2, value=old_passed_count)
    ws.cell(row=summary_metric_row, column=3, value="旧版不合格")
    ws.cell(row=summary_metric_row, column=4, value=old_failed_count)
    ws.cell(row=summary_metric_row, column=5, value="修复")
    ws.cell(row=summary_metric_row, column=6, value=fixed_count)
    ws.cell(row=summary_metric_row, column=7, value="新增违规")
    ws.cell(row=summary_metric_row, column=8, value=regressed_count)
    ws.cell(row=summary_metric_row, column=9, value="新版合格")
    ws.cell(row=summary_metric_row, column=10, value=new_passed_count)
    ws.cell(row=summary_metric_row, column=11, value="新版不合格")
    ws.cell(row=summary_metric_row, column=12, value=new_failed_count)
    ws.cell(row=summary_problem_row, column=1, value="问题定位")
    if problem_items:
        ws.cell(row=summary_problem_row, column=2, value="；".join(problem_items))
    else:
        ws.cell(row=summary_problem_row, column=2, value="无不合格或异常")

    # 格式化摘要行
    for row in range(1, summary_rows + 1 + persona_row_offset):
        r = row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col)
            cell.alignment = wrap_align
            cell.border = thin_border
            if has_persona_header and r == 1:
                continue  # 已设置
            elif r == summary_title_row:
                cell.fill = summary_title_fill
                cell.font = summary_title_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif r == summary_metric_row:
                cell.fill = summary_metric_fill
                if col in (1, 3, 5, 7, 9, 11):
                    cell.font = summary_label_font
                else:
                    cell.font = cell_font
            else:
                cell.fill = summary_problem_fill if problem_items else pass_fill
                if col == 1:
                    cell.font = summary_label_font
                else:
                    cell.font = cell_font

    ws.merge_cells(start_row=summary_title_row, start_column=1, end_row=summary_title_row, end_column=len(headers))
    ws.merge_cells(start_row=summary_problem_row, start_column=2, end_row=summary_problem_row, end_column=len(headers))
    ws.row_dimensions[summary_title_row].height = 24
    ws.row_dimensions[summary_metric_row].height = 22
    ws.row_dimensions[summary_problem_row].height = 42

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    old_map = {r["category"]: r for r in results_old}
    new_map = {r["category"]: r for r in results_new}
    all_categories = list(dict.fromkeys([r["category"] for r in results_old + results_new]))

    for i, cat in enumerate(all_categories):
        row = data_start_row + i
        old_r = old_map.get(cat, {})
        new_r = new_map.get(cat, {})

        old_passed = old_r.get("passed", False)
        new_passed = new_r.get("passed", False)

        if old_passed is None or new_passed is None:
            change = "跳过质检"
            change_fill = None
        elif old_passed and new_passed:
            change = "-"
            change_fill = None
        elif not old_passed and not new_passed:
            change = "仍不合格"
            change_fill = fail_fill
        elif not old_passed and new_passed:
            change = "✓ 修复"
            change_fill = improve_fill
        else:
            change = "✗ 新增违规"
            change_fill = regress_fill

        detail = ""
        if change in ["✓ 修复", "✗ 新增违规"]:
            old_summary = old_r.get("violation_summary", "")
            new_summary = new_r.get("violation_summary", "")
            if old_summary or new_summary:
                detail = f"旧版: {old_summary}\n新版: {new_summary}"

        values = [
            i + 1,
            cat,
            old_r.get("dialogue", ""),
            old_r.get("ticket_json", ""),
            "跳过质检" if old_passed is None else ("合格" if old_passed else ("不合格" if not old_r.get("error") else "异常")),
            old_r.get("violation_summary", ""),
            new_r.get("dialogue", ""),
            new_r.get("ticket_json", ""),
            "跳过质检" if new_passed is None else ("合格" if new_passed else ("不合格" if not new_r.get("error") else "异常")),
            new_r.get("violation_summary", ""),
            change,
            detail,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = cell_font
            cell.alignment = wrap_align
            cell.border = thin_border

        old_result_cell = ws.cell(row=row, column=5)
        new_result_cell = ws.cell(row=row, column=9)
        if old_passed is True:
            old_result_cell.fill = pass_fill
        elif old_passed is False and not old_r.get("error"):
            old_result_cell.fill = fail_fill
        if new_passed is True:
            new_result_cell.fill = pass_fill
        elif new_passed is False and not new_r.get("error"):
            new_result_cell.fill = fail_fill

        change_cell = ws.cell(row=row, column=11)
        if change_fill:
            change_cell.fill = change_fill

    col_widths = [6, 14, 50, 40, 10, 40, 50, 40, 10, 40, 12, 60]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = f"A{data_start_row}"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{header_row}:{last_col}{data_start_row + len(all_categories) - 1}"


def write_regression_excel(results_old, results_new, output_file, old_a_ver, new_a_ver, old_b_ver, new_b_ver):
    """回归对比：单C人设输出一个Excel（兼容旧接口）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "回归对比"
    write_regression_sheet(ws, results_old, results_new, old_a_ver, new_a_ver, old_b_ver, new_b_ver)
    wb.save(output_file)
    print(f"\n回归对比结果已保存到: {output_file}")


def write_multi_c_regression_excel(results_by_c, output_file, old_a_ver, new_a_ver, old_b_ver, new_b_ver,
                                   scenario_order, persona_c_desc_map=None, optimization_data=None):
    """多C人设回归对比：每个C人设一个独立工作表。optimization_data: (analysis_list, revised_persona) 或 None。"""
    persona_c_desc_map = persona_c_desc_map or {}
    wb = openpyxl.Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    for c_ver, (results_old, results_new) in results_by_c.items():
        label = short_persona_label(persona_c_desc_map.get(c_ver, ""))
        ws = wb.create_sheet((label or c_ver)[:31])
        old_ordered = sorted(
            results_old,
            key=lambda r: scenario_order.index(r["category"]) if r["category"] in scenario_order else 999,
        )
        new_ordered = sorted(
            results_new,
            key=lambda r: scenario_order.index(r["category"]) if r["category"] in scenario_order else 999,
        )
        write_regression_sheet(
            ws, old_ordered, new_ordered, old_a_ver, new_a_ver, old_b_ver, new_b_ver,
            c_persona_label=c_ver,
            c_persona_desc=persona_c_desc_map.get(c_ver, ""),
        )

    if optimization_data:
        analysis_list, revised_persona = optimization_data
        ws_opt = wb.create_sheet("优化建议")
        write_optimization_sheet(ws_opt, analysis_list, revised_persona, new_a_ver)

    wb.save(output_file)
    print(f"\n多C回归对比结果已保存到: {output_file}")


def main():
    global API_KEY, BASE_URL, MODEL_A, MODEL_B, MODEL_C, MODEL_D
    global CONCURRENCY, PERSONA_CONCURRENCY, API_CONCURRENCY, MAX_ROUNDS, MAX_RETRIES, RETRY_DELAY, SCENARIO_RETRIES, API_SEMAPHORE

    parser = argparse.ArgumentParser(description="AI外呼用户反应测试脚本")
    parser.add_argument("--config", default=DEFAULT_CONFIG_FILE,
                        help="配置文件路径，默认读取config.local.json")
    parser.add_argument("--persona", nargs=2, metavar=("A_VER", "B_VER"),
                        help="指定人设版本，如: --persona A_v2 B_v1")
    parser.add_argument("--regression", action="store_true",
                        help="启用回归对比模式")
    parser.add_argument("--old", nargs=2, metavar=("A_VER", "B_VER"),
                        help="回归对比的旧版本，如: --old A_v1 B_v1")
    parser.add_argument("--new", nargs=2, metavar=("A_VER", "B_VER"),
                        help="回归对比的新版本，如: --new A_v2 B_v1")
    parser.add_argument("--persona-c", default=None,
                        help="C模型人设版本，如: C_v1(普通) C_v2(投诉) C_v3(法律) C_v4(诱导) C_v5(情绪)")
    parser.add_argument("--persona-c-list", nargs="+",
                        help="一次运行多个C模型人设，并为每个C人设生成独立工作表，如: --persona-c-list C_v1 C_v2 C_v3 C_v4 C_v5")
    parser.add_argument("--persona-d", default=None,
                        help="D模型(工单生成)人设版本，如: D_v1")
    parser.add_argument("--limit", type=int, default=0,
                        help="限制测试场景数量，0=全部")
    parser.add_argument("--sample-size", type=int, default=0,
                        help="随机抽样测试的场景分类数量，0=不抽样")
    parser.add_argument("--sample-seed", type=int, default=42,
                        help="随机抽样种子，便于复现")
    parser.add_argument("--concurrency", type=int, default=None,
                        help=f"并发场景数，默认读取TEST_CONCURRENCY或{CONCURRENCY}")
    parser.add_argument("--persona-concurrency", type=int, default=None,
                        help=f"多C人设并发数，默认读取PERSONA_CONCURRENCY或{PERSONA_CONCURRENCY}")
    parser.add_argument("--api-concurrency", type=int, default=None,
                        help=f"全局API请求并发上限，默认读取API_CONCURRENCY或{API_CONCURRENCY}")
    parser.add_argument("--max-rounds", type=int, default=None,
                        help=f"单通对话最大轮次，默认读取MAX_ROUNDS或{MAX_ROUNDS}")
    parser.add_argument("--max-retries", type=int, default=None,
                        help=f"模型调用失败最大重试次数，默认读取MAX_RETRIES或{MAX_RETRIES}")
    parser.add_argument("--scenario-retries", type=int, default=None,
                        help=f"场景失败后完整重跑次数，默认读取SCENARIO_RETRIES或{SCENARIO_RETRIES}")
    parser.add_argument("--skip-audit", action="store_true",
                        help="只生成A/C对话，不调用B模型质检，用于快速冒烟测试")
    parser.add_argument("--skip-ticket", action="store_true",
                        help="跳过D模型工单生成；与--skip-audit互相独立")
    parser.add_argument("--analyze-failures", action="store_true",
                        help="对不合格场景调用模型分析A人设根因并生成优化建议")

    args = parser.parse_args()

    config = load_config(args.config)
    if config:
        API_KEY = config.get("api_key") or API_KEY
        BASE_URL = config.get("base_url") or BASE_URL
        MODEL_A = config.get("model_a") or MODEL_A
        MODEL_B = config.get("model_b") or MODEL_B
        MODEL_C = config.get("model_c") or MODEL_C
        MODEL_D = config.get("model_d") or MODEL_D
        CONCURRENCY = int(config.get("concurrency", CONCURRENCY))
        PERSONA_CONCURRENCY = int(config.get("persona_concurrency", PERSONA_CONCURRENCY))
        API_CONCURRENCY = int(config.get("api_concurrency", API_CONCURRENCY))
        MAX_ROUNDS = int(config.get("max_rounds", MAX_ROUNDS))
        MAX_RETRIES = int(config.get("max_retries", MAX_RETRIES))
        RETRY_DELAY = int(config.get("retry_delay", RETRY_DELAY))
        SCENARIO_RETRIES = int(config.get("scenario_retries", SCENARIO_RETRIES))
    if args.concurrency is not None:
        CONCURRENCY = args.concurrency
    if args.persona_concurrency is not None:
        PERSONA_CONCURRENCY = args.persona_concurrency
    if args.api_concurrency is not None:
        API_CONCURRENCY = args.api_concurrency
    if args.max_rounds is not None:
        MAX_ROUNDS = args.max_rounds
    if args.max_retries is not None:
        MAX_RETRIES = args.max_retries
    if args.scenario_retries is not None:
        SCENARIO_RETRIES = args.scenario_retries
    API_CONCURRENCY = max(1, API_CONCURRENCY)
    API_SEMAPHORE = threading.BoundedSemaphore(API_CONCURRENCY)

    if not API_KEY:
        print("  [错误] 未设置api_key。请在config.local.json中配置api_key，或设置ARK_API_KEY环境变量")
        sys.exit(1)

    print("=" * 60)
    print("  AI外呼用户反应测试 v3.0")
    print(f"  A模型(AI客服): {MODEL_A}")
    print(f"  B模型(质检):   {MODEL_B}")
    print(f"  C模型(用户):   {MODEL_C}")
    print(f"  D模型(工单):   {MODEL_D}")
    task_peak = CONCURRENCY * PERSONA_CONCURRENCY
    print(f"  场景并发数: {CONCURRENCY}  C人设并发数: {PERSONA_CONCURRENCY}  最大轮次: {MAX_ROUNDS}")
    print(f"  估算任务峰值: {task_peak}  全局API并发上限: {API_CONCURRENCY}")
    if task_peak > API_CONCURRENCY:
        print(f"  [提示] 任务峰值超过API并发上限，脚本会自动排队限流；如需更稳，可降低 --concurrency 或 --persona-concurrency。")
    if API_CONCURRENCY > 60:
        print("  [提示] API并发上限较高，可能触发连接错误或平台限流；建议全量测试先使用 20-40。")
    if SCENARIO_RETRIES:
        print(f"  场景级重试: {SCENARIO_RETRIES} 次，仅重试连接/限流/超时/5xx类异常")
    print("=" * 60)
    total_script_start = time.time()

    # 加载场景
    scenarios = select_scenarios(DEFAULT_SCENARIOS, args.limit, args.sample_size, args.sample_seed)
    print(f"\n[1/3] 场景列表：共 {len(DEFAULT_SCENARIOS)} 个，本次运行 {len(scenarios)} 个")
    if args.sample_size > 0:
        print(f"  随机抽样 {len(scenarios)} 个分类，seed={args.sample_seed}")
    elif args.limit > 0:
        print(f"  限制测试前 {len(scenarios)} 个分类")

    try:
        from openai import OpenAI
    except ImportError:
        print("  [错误] 未安装openai依赖，请先执行: pip install -r requirements.txt")
        sys.exit(1)

    client = OpenAI(base_url=BASE_URL, api_key=API_KEY)

    # 解析 D 人设版本：命令行 > 配置文件 > 自动检测最新
    if args.persona_d:
        d_ver = args.persona_d
    elif config.get("persona_d"):
        d_ver = config["persona_d"]
    else:
        try:
            d_ver = get_latest_version("D")
        except FileNotFoundError:
            d_ver = None

    persona_d = None
    if d_ver and not args.skip_ticket:
        try:
            persona_d = load_persona(d_ver)
            print(f"  D人设: {d_ver} 已加载")
        except FileNotFoundError as e:
            print(f"  [警告] D人设加载失败，将跳过工单生成: {e}")
            d_ver = None
    elif args.skip_ticket:
        print("  [提示] --skip-ticket 已开启，本次不生成工单")

    if args.regression:
        # 回归对比模式
        if not args.old or not args.new:
            print("  [错误] 回归模式需要 --old 和 --new 参数")
            sys.exit(1)

        old_a_ver, old_b_ver = args.old
        new_a_ver, new_b_ver = args.new

        # 解析 C 人设列表：--persona-c-list > --persona-c > 配置文件 > 默认
        if args.persona_c_list:
            c_versions = args.persona_c_list
        elif args.persona_c:
            c_versions = [args.persona_c]
        elif config.get("persona_c_list"):
            c_versions = config["persona_c_list"]
        else:
            c_versions = [config.get("persona_c") or "C_v1"]

        c_list_str = ",".join(c_versions)
        print(f"\n[2/3] 回归对比模式: {old_a_ver}/{old_b_ver} vs {new_a_ver}/{new_b_ver}, C={c_list_str}")

        # 加载 A/B 人设
        try:
            persona_a_old = load_persona(old_a_ver)
            persona_b_old = load_persona(old_b_ver)
            persona_a_new = load_persona(new_a_ver)
            persona_b_new = load_persona(new_b_ver)
        except FileNotFoundError as e:
            print(f"  [错误] {e}")
            sys.exit(1)

        # 加载所有 C 人设
        persona_c_map = {}
        persona_c_desc_map = {}
        for c_ver in c_versions:
            try:
                content = load_persona(c_ver)
                persona_c_map[c_ver] = content
                persona_c_desc_map[c_ver] = extract_persona_desc(content)
            except FileNotFoundError as e:
                print(f"  [错误] {e}")
                sys.exit(1)

        # 逐个 C 人设跑回归对比
        results_by_c = {}
        elapsed_by_c = {}
        total_start = time.time()
        for c_ver in c_versions:
            persona_c = persona_c_map[c_ver]
            print(f"\n  [{c_ver}] 并发执行新旧版本测试...")
            c_start = time.time()
            with ThreadPoolExecutor(max_workers=2) as executor:
                future_old = executor.submit(run_test, client, scenarios, persona_a_old, persona_b_old, persona_c,
                                             "旧版", args.skip_audit, persona_d, args.skip_ticket)
                future_new = executor.submit(run_test, client, scenarios, persona_a_new, persona_b_new, persona_c,
                                             "新版", args.skip_audit, persona_d, args.skip_ticket)
                results_old, elapsed_old = future_old.result()
                results_new, elapsed_new = future_new.result()
            c_elapsed = time.time() - c_start

            results_by_c[c_ver] = (results_old, results_new)
            elapsed_by_c[c_ver] = (elapsed_old, elapsed_new)

            old_passed = sum(1 for r in results_old if r["passed"] and not r["error"])
            new_passed = sum(1 for r in results_new if r["passed"] and not r["error"])
            fixed = sum(1 for r_old, r_new in zip(results_old, results_new) if not r_old["passed"] and r_new["passed"])
            regressed = sum(1 for r_old, r_new in zip(results_old, results_new) if r_old["passed"] and not r_new["passed"])
            print(f"  [{c_ver}] 旧版: {old_passed}/{len(results_old)}  新版: {new_passed}/{len(results_new)}  "
                  f"修复: {fixed}  新增违规: {regressed}  耗时: {format_duration(c_elapsed)}")

        total_elapsed = time.time() - total_start

        # 优化分析（基于新版A人设的新版结果）
        opt_data = None
        if args.analyze_failures:
            print(f"\n  调用优化分析模型...")
            all_failed_new = []
            for _, results_new in results_by_c.values():
                all_failed_new.extend([r for r in results_new if r.get("passed") is False and not r.get("error")])
            violations_text = collect_violations(all_failed_new)
            if violations_text:
                opt_data = generate_optimization(client, persona_a_new, violations_text)
                if opt_data[0]:
                    print(f"  优化建议已生成 ({len(opt_data[0])} 条)")
                else:
                    print(f"  优化分析未产出结果")
            else:
                print(f"  无不合格项，跳过优化分析")

        # 写 Excel
        print(f"\n[3/3] 写入多C回归对比Excel...")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        c_part = "-".join(c_versions)
        latest_name = f"回归对比_{old_a_ver}_vs_{new_a_ver}_{timestamp}.xlsx"
        output_file, latest_file = build_output_paths(
            timestamp,
            f"regression_{old_a_ver}_{old_b_ver}_vs_{new_a_ver}_{new_b_ver}_{c_part}",
            latest_name,
        )
        cat_order = list(scenarios)
        write_multi_c_regression_excel(results_by_c, output_file, old_a_ver, new_a_ver, old_b_ver, new_b_ver,
                                       cat_order, persona_c_desc_map, opt_data)
        copy_to_latest(output_file, latest_file)

        # 更新变更日志（为每个 C 人设写入一条）
        for c_ver in c_versions:
            results_old, results_new = results_by_c[c_ver]
            update_changelog(old_a_ver, new_a_ver, old_b_ver, new_b_ver, results_old, results_new)

        # 汇总
        print("\n" + "=" * 60)
        print(f"  回归对比完成: {old_a_ver}/{old_b_ver} vs {new_a_ver}/{new_b_ver}")
        for c_ver in c_versions:
            results_old, results_new = results_by_c[c_ver]
            old_p = sum(1 for r in results_old if r["passed"] and not r["error"])
            new_p = sum(1 for r in results_new if r["passed"] and not r["error"])
            el_old, el_new = elapsed_by_c[c_ver]
            print(f"  {c_ver}: {old_p}/{len(results_old)} → {new_p}/{len(results_new)}  旧耗时: {format_duration(el_old)}  新耗时: {format_duration(el_new)}")
        print(f"  总耗时: {format_duration(total_elapsed)}")
        print(f"  本次测试总耗时: {format_duration(time.time() - total_script_start)}")
        print("=" * 60)

    else:
        # 单版本测试模式
        if args.persona:
            a_ver, b_ver = args.persona
        else:
            a_ver = config.get("persona_a") or get_latest_version("A")
            b_ver = config.get("persona_b") or get_latest_version("B")

        if args.persona_c_list:
            c_versions = args.persona_c_list
        elif args.persona_c:
            c_versions = [args.persona_c]
        elif config.get("persona_c_list"):
            c_versions = config["persona_c_list"]
        else:
            c_versions = [config.get("persona_c") or "C_v1"]
        print(f"\n[2/3] 单版本测试: A={a_ver}, B={b_ver}, C={','.join(c_versions)}")

        try:
            persona_a = load_persona(a_ver)
            persona_b = load_persona(b_ver)
            persona_c_map = {c_ver: load_persona(c_ver) for c_ver in c_versions}
        except FileNotFoundError as e:
            print(f"  [错误] {e}")
            sys.exit(1)

        persona_c_desc_map = {c_ver: extract_persona_desc(content) for c_ver, content in persona_c_map.items()}

        cat_order = list(scenarios)

        if len(c_versions) > 1:
            results_by_c, elapsed_map = run_test_for_personas(
                client, scenarios, persona_a, persona_b, persona_c_map,
                args.skip_audit, PERSONA_CONCURRENCY,
                persona_d, args.skip_ticket,
            )
            for results in results_by_c.values():
                results.sort(key=lambda r: cat_order.index(r["category"]) if r["category"] in cat_order else 999)

            # 优化分析
            opt_data = None
            if args.analyze_failures:
                print(f"\n  调用优化分析模型...")
                all_failed = []
                for results in results_by_c.values():
                    all_failed.extend([r for r in results if r.get("passed") is False and not r.get("error")])
                violations_text = collect_violations(all_failed)
                if violations_text:
                    opt_data = generate_optimization(client, persona_a, violations_text)
                    if opt_data[0]:
                        print(f"  优化建议已生成 ({len(opt_data[0])} 条)")
                    else:
                        print(f"  优化分析未产出结果")
                else:
                    print(f"  无不合格项，跳过优化分析")

            print(f"\n[3/3] 写入多C人设Excel...")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            latest_name = f"AI外呼多C测试结果_{a_ver}_{b_ver}_{timestamp}.xlsx"
            output_file, latest_file = build_output_paths(
                timestamp,
                f"{a_ver}_{b_ver}_{'-'.join(c_versions)}",
                latest_name,
            )
            write_multi_c_excel(results_by_c, output_file, a_ver, b_ver, cat_order, persona_c_desc_map, opt_data)
            copy_to_latest(output_file, latest_file)

            print("\n" + "=" * 60)
            print(f"  多C人设测试完成: 共{len(scenarios)}个分类，{len(c_versions)}个C人设")
            for c_ver in c_versions:
                results = results_by_c[c_ver]
                passed_count = sum(1 for r in results if r["passed"] and not r["error"])
                skipped_count = sum(1 for r in results if r["passed"] is None and not r["error"])
                failed_count = sum(1 for r in results if r["passed"] is False and not r["error"])
                error_count = sum(1 for r in results if r["error"])
                print(f"  {c_ver}: 合格 {passed_count}  不合格 {failed_count}  跳过质检 {skipped_count}  异常 {error_count}  耗时 {format_duration(elapsed_map[c_ver])}")
            print(f"  本次测试总耗时: {format_duration(time.time() - total_script_start)}")
            print("=" * 60)
            return

        persona_c = next(iter(persona_c_map.values()))
        results, elapsed = run_test(client, scenarios, persona_a, persona_b, persona_c,
                                    skip_audit=args.skip_audit,
                                    persona_d=persona_d, skip_ticket=args.skip_ticket)
        results.sort(key=lambda r: cat_order.index(r["category"]) if r["category"] in cat_order else 999)

        # 优化分析
        opt_data = None
        if args.analyze_failures:
            print(f"\n  调用优化分析模型...")
            failed = [r for r in results if r.get("passed") is False and not r.get("error")]
            violations_text = collect_violations(failed)
            if violations_text:
                opt_data = generate_optimization(client, persona_a, violations_text)
                if opt_data[0]:
                    print(f"  优化建议已生成 ({len(opt_data[0])} 条)")
                else:
                    print(f"  优化分析未产出结果")
            else:
                print(f"  无不合格项，跳过优化分析")

        print(f"\n[3/3] 写入Excel...")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        c_ver = next(iter(persona_c_map.keys()))
        latest_name = f"AI外呼测试结果_{a_ver}_{b_ver}_{c_ver}_{timestamp}.xlsx"
        output_file, latest_file = build_output_paths(
            timestamp,
            f"{a_ver}_{b_ver}_{c_ver}",
            latest_name,
        )
        write_single_excel(results, output_file, a_ver, b_ver,
                           c_persona_label=c_ver,
                           c_persona_desc=persona_c_desc_map.get(c_ver, ""),
                           optimization_data=opt_data)
        copy_to_latest(output_file, latest_file)

        total_scenarios = len(results)
        passed_count = sum(1 for r in results if r["passed"] and not r["error"])
        skipped_count = sum(1 for r in results if r["passed"] is None and not r["error"])
        failed_count = sum(1 for r in results if r["passed"] is False and not r["error"])
        error_count = sum(1 for r in results if r["error"])

        print("\n" + "=" * 60)
        print(f"  测试完成: 共{total_scenarios}个分类")
        print(f"  合格: {passed_count}  不合格: {failed_count}  跳过质检: {skipped_count}  异常: {error_count}")
        print(f"  总耗时: {format_duration(elapsed)}")
        print(f"  本次测试总耗时: {format_duration(time.time() - total_script_start)}")
        print(f"  平均每个场景: {format_duration(elapsed / total_scenarios)}")
        print("=" * 60)


if __name__ == "__main__":
    main()
